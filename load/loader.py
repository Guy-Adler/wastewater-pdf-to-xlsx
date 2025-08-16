"""
Basic loader from structured data to excel.

Assumptions made (should be kept to a minimum):
- There are at least 2 rows of data
- The second row of data is good to use as a "format template"
- If there are rows in which the first column is not a date, they are at the end of the row list.
- Date rows are continuos (there are no breaks of non date values in the date column)
"""

import datetime
import sys
from typing import Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import openpyxl.utils
from copy import copy
from .schemas import LoadSchemaManager
from .utils import WorkbookContext

def column_index_from_string(col_letter: str) -> int:
  return openpyxl.utils.column_index_from_string(col_letter) - 1

def extract_date_from_row(r: tuple[Cell], date_column: int) -> datetime.date | None:
  if type(r[date_column].value) is datetime.date:
    return r[date_column].value
  if type(r[date_column].value) is datetime.datetime:
    return r[date_column].value.date()
  return None

schemaManager = LoadSchemaManager()

class Loader:

  file_path: str
  workbook: Workbook
  worksheet: Worksheet
  schema: dict
  sheet_schema: Any

  def __init__(self, file_path: str, schema_name: str):
    self.schema = schemaManager.get_schema(schema_name)
    if not self.schema:
      raise ValueError(f"No schema found for name: {schema_name}")
    self.file_path = file_path

  def _get_max_date_row(self):
    date_column = column_index_from_string(self.sheet_schema.get("fields", {}).get("date", {}).get("column", "A"))
    row_idx = self.worksheet.max_row
    first_row = self.sheet_schema.get('headerRowCount', 0) + 1
    while row_idx >= first_row:
      if extract_date_from_row(self.worksheet[row_idx], date_column) is not None:
        return row_idx
      row_idx -= 1
    raise Exception('Could not find a date row')

  def _get_row(self, date: datetime.date):
    """
    Binary search for the row with the given date in the first column.
    If not found, insert a new row in the correct place (sorted by date).
    Also insert all rows between the missing date and the previous/next dates.
    Returns the row of the given date.
    """
    date_column = column_index_from_string(self.sheet_schema.get("fields", {}).get("date", {}).get("column", "A"))
    min_date_row = self.sheet_schema.get('headerRowCount', 0) + 1
    max_date_row = self._get_max_date_row()

    start = min_date_row
    end = max_date_row

    while start <= end:
      mid = (start + end) // 2
      row = self.worksheet[mid]
      row_date = extract_date_from_row(row, date_column)
      if row_date is None:
        end -= 1
        continue

      if date == row_date:
        return row
      elif row_date < date:
        start = mid + 1
      else:
        end = mid - 1

    # Insert missing rows
    one_day = datetime.timedelta(days=1)
    start_date = date - one_day if end < min_date_row else extract_date_from_row(self.worksheet[end], date_column)
    end_date = date if start >= max_date_row else extract_date_from_row(self.worksheet[start], date_column) - one_day

    current_date = end_date
    date_row: tuple[Cell] = None

    # Fill all rows in range (start_date, end_date]
    while current_date > start_date:
      row = self._add_row(start)

      if current_date == date:
        date_row = row

      row[date_column].value = current_date
      current_date -= one_day

    return date_row
  
  def _add_row(self, row_index: int, template_row: int = None):
    """
    Add an empty row at the specified index (push all other rows 1 down).
    Copy styles from the template row (default the first row after the headers).
    """
    if template_row is None:
      template_row = self.sheet_schema.get('headerRowCount', 0) + 2 # use the second row after the header as a template
    self.worksheet.insert_rows(row_index)

    # Copy each cell's style
    for col, src_cell in enumerate(self.worksheet[template_row + 1], start=1): # need to account for the newly added row
      tgt_cell = self.worksheet.cell(row=row_index, column=col)
      if src_cell.has_style:
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)

    # Copy row height
    if self.worksheet.row_dimensions[template_row].height is not None:
      self.worksheet.row_dimensions[row_index].height = self.worksheet.row_dimensions[template_row].height

    # Copy merged cells (same row only)
    for merged_range in list(self.worksheet.merged_cells.ranges):
        min_row, min_col, max_row, max_col = merged_range.bounds
        if min_row == template_row and max_row == template_row:
            new_range = (
                self.worksheet.cell(row=row_index, column=min_col).coordinate
                + ":"
                + self.worksheet.cell(row=row_index, column=max_col).coordinate
            )
            self.worksheet.merge_cells(new_range)

    return self.worksheet[row_index]
  
  def load(self, data: dict):
    self.sheet_schema = self.schema.get('sheets', {}).get(data['type'])
    if not self.sheet_schema:
      raise ValueError(f"No sheet defined for type: {data['type']}")
    
    with WorkbookContext(self.file_path) as self.workbook:
      self.worksheet = self.workbook[self.sheet_schema['name']]
      row = self._get_row(data['sampling_date'].date())
      if not row:
        print(f'No existing row found for date {data["sampling_date"]}', file=sys.stderr)
        return

      for field, field_schema in self.sheet_schema.get('fields', {}).items():
        if field in data["results"]:
          col_index = column_index_from_string(field_schema['column'])
          row[col_index].value = data["results"][field]
